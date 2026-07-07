import argparse
import json
import re
from pathlib import Path


BAD_TERMS = ["hostel", "hostal", "camping", "camper", "camps"]
APARTMENT_TERMS = ["apartment", "apartamento", "apartamentos", "apartament", "studio"]


def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def normalize_url(value):
    return str(value or "").strip()


def is_hotel_like(hotel):
    name = (hotel.get("hotel_name") or hotel.get("name") or "").lower()
    property_type = (hotel.get("property_type") or "").lower()
    if any(term in name for term in BAD_TERMS):
        return False
    return property_type == "hotel" or "hotel" in name


def accommodation_label(hotel):
    name = (hotel.get("hotel_name") or hotel.get("name") or "").lower()
    property_type = (hotel.get("property_type") or "").lower()
    if property_type == "hotel" or "hotel" in name:
        return "hotel"
    if any(term in name for term in APARTMENT_TERMS):
        return "apartament"
    return "apartament"


def destination_label(flight, hotel):
    label = flight.get("dest_name") or flight.get("stay_city") or flight.get("dest_iata") or ""
    return f"{label} / {accommodation_label(hotel)}"


def flatten_hotel_results(paths):
    rows = []
    for source in paths:
        data = load_json(source)
        for batch in data:
            flight = batch.get("flight") or {}
            for hotel in batch.get("hotels") or []:
                hotel_price = hotel.get("hotel_price_per_person")
                flight_price = flight.get("flight_price")
                if hotel_price is None or flight_price is None:
                    continue
                total = round(float(hotel_price) + float(flight_price), 2)
                rows.append({
                    "source_file": str(source),
                    "origin": flight.get("origin"),
                    "origin_iata": flight.get("origin_iata"),
                    "country": flight.get("country"),
                    "dest_iata": flight.get("dest_iata"),
                    "dest_name": flight.get("dest_name"),
                    "stay_city": flight.get("stay_city"),
                    "depart": flight.get("depart"),
                    "return": flight.get("return"),
                    "days": int(flight.get("days") or 0),
                    "nights": int(flight.get("nights") or 0),
                    "flight_price": round(float(flight_price), 2),
                    "flight_link": flight.get("flight_link"),
                    "hotel_name": hotel.get("hotel_name"),
                    "hotel_area": hotel.get("hotel_area"),
                    "property_type": hotel.get("property_type"),
                    "accommodation_type": hotel.get("accommodation_type"),
                    "stars": hotel.get("stars"),
                    "review_score": hotel.get("review_score"),
                    "review_count": hotel.get("review_count"),
                    "hotel_price": round(float(hotel_price), 2),
                    "hotel_link": hotel.get("hotel_link"),
                    "total": total,
                })
    return rows


def load_pool_map(path):
    if not path:
        return {}
    data = load_json(path)
    result = {}
    for item in data:
        link = normalize_url(item.get("hotel_link"))
        if not link:
            continue
        result[link] = {
            "has_outdoor_pool": item.get("has_outdoor_pool"),
            "outdoor_evidence": item.get("outdoor_evidence") or "",
        }
    return result


def apply_filters(rows, args, pool_map):
    seen = set()
    output = []
    dest_filter = {value.strip().upper() for value in args.dest} if args.dest else set()
    for row in rows:
        key = (row.get("origin_iata"), row.get("dest_iata"), row.get("depart"), row.get("return"), normalize_url(row.get("hotel_link")))
        if key in seen:
            continue
        seen.add(key)

        if args.max_total is not None and row["total"] > args.max_total:
            continue
        if args.min_total is not None and row["total"] < args.min_total:
            continue
        if args.return_date and row.get("return") != args.return_date:
            continue
        if args.depart_from and row.get("depart") < args.depart_from:
            continue
        if args.depart_to and row.get("depart") > args.depart_to:
            continue
        if args.min_days and row["days"] < args.min_days:
            continue
        if dest_filter and (row.get("dest_iata") or "").upper() not in dest_filter:
            continue
        if args.hotel_only and not is_hotel_like(row):
            continue

        pool = pool_map.get(normalize_url(row.get("hotel_link")), {})
        row["has_outdoor_pool"] = pool.get("has_outdoor_pool")
        row["outdoor_evidence"] = pool.get("outdoor_evidence", "")
        if args.confirmed_outdoor_only and row["has_outdoor_pool"] is not True:
            continue
        output.append(row)

    output.sort(key=lambda item: (
        item["total"],
        -(float(item.get("review_score") or 0)),
        -(int(item.get("review_count") or 0)),
    ))
    if args.limit:
        output = output[:args.limit]
    return output


def format_pln(value):
    return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", " ") + f" z{chr(0x0142)}"


def write_xlsx(rows, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("Install openpyxl first: python -m pip install -r requirements.txt") from exc

    headers = [
        "Miasto wylot",
        "Kraj",
        "Miasto dolecowe",
        "Lot cena",
        "Lot link",
        "Hotel cena",
        "Hotel link",
        "Basen",
        "Cena total",
        "Ilosc dni wyjazdu",
        "Review score",
        "Review count",
        "Hotel area",
        "Evidence",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Offers"
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num, item in enumerate(rows, start=2):
        if item.get("has_outdoor_pool") is True:
            pool_label = "TAK"
        elif item.get("has_outdoor_pool") is False:
            pool_label = "NIE/KRYTY"
        else:
            pool_label = "NIEPOTW."
        values = [
            item.get("origin"),
            item.get("country") or "Hiszpania",
            destination_label(item, item),
            format_pln(item.get("flight_price")),
            "Lot",
            format_pln(item.get("hotel_price")),
            item.get("hotel_name"),
            pool_label,
            format_pln(item.get("total")),
            item.get("days"),
            item.get("review_score"),
            item.get("review_count"),
            item.get("hotel_area"),
            item.get("outdoor_evidence"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_num, col, value)
            ws.cell(row_num, col).alignment = Alignment(vertical="top", wrap_text=True)
        if item.get("flight_link"):
            ws.cell(row_num, 5).hyperlink = item.get("flight_link")
            ws.cell(row_num, 5).style = "Hyperlink"
        if item.get("hotel_link"):
            ws.cell(row_num, 7).hyperlink = item.get("hotel_link")
            ws.cell(row_num, 7).style = "Hyperlink"

    widths = [18, 16, 34, 14, 12, 14, 48, 12, 14, 18, 12, 12, 24, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(ws.max_row, 1)}"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Build deterministic offer list from Agoda hotel result JSON files.")
    parser.add_argument("--hotel-results", nargs="+", required=True, help="Agoda result JSON files from agoda_search.js")
    parser.add_argument("--pool-check", help="JSON from verify_outdoor_pool.js")
    parser.add_argument("--out-json", default="output/offers.json")
    parser.add_argument("--out-xlsx")
    parser.add_argument("--max-total", type=float)
    parser.add_argument("--min-total", type=float)
    parser.add_argument("--return-date")
    parser.add_argument("--depart-from")
    parser.add_argument("--depart-to")
    parser.add_argument("--min-days", type=int, default=0)
    parser.add_argument("--dest", nargs="*", help="Destination IATA filters, e.g. AGP VLC ALC")
    parser.add_argument("--hotel-only", action="store_true")
    parser.add_argument("--confirmed-outdoor-only", action="store_true")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    rows = flatten_hotel_results(args.hotel_results)
    pool_map = load_pool_map(args.pool_check)
    offers = apply_filters(rows, args, pool_map)

    out_json = Path(args.out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(offers, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"offers: {len(offers)} -> {out_json}")

    if args.out_xlsx:
        write_xlsx(offers, args.out_xlsx)
        print(f"xlsx -> {args.out_xlsx}")


if __name__ == "__main__":
    main()
